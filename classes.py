from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image

class LeitorArquivos:
    
    def __init__(self, caminho_arquivo: str = ""):
        self.caminho_arquivo = caminho_arquivo
        self.dados = []
        
        
    def processa_arquivo(self, acao:str):
        with open(f'{self.caminho_arquivo}{acao}.txt','r') as arq_cotacao:
            linhas = arq_cotacao.readlines()
            self.dados = [linha.replace('\n',"").split(';') for linha in linhas]
            

class PropriedadeSeries:
    
    def __init__(self, espessura: int, cor: str):
        self.espessura = espessura
        self.cor = cor

            
class GerenciadorPlanilhas:
    
    def __init__(self):
        self.workbook = Workbook()
        self.ws_ativa = None
        
        
    def adiciona_planilha(self, titulo:str):
        nova_planilha = self.workbook.create_sheet(titulo)
        self.workbook.active = nova_planilha
        self.ws_ativa = nova_planilha
        
        return nova_planilha
    
    
    def adiciona_linha(self, dados: list):
        self.ws_ativa.append(dados)
        
        
    def atualiza_celula(self, celula: str, dado):
        self.ws_ativa[celula] = dado
        
        
    def mescla_celulas(self, cel_inicio: str, cel_fim: str):
        self.ws_ativa.merge_cells(f'{cel_inicio}:{cel_fim}')
        
        
    def aplica_estilos(self, celula:str, estilos: list):
        
        for estilo in estilos:
            setattr(self.ws_ativa[celula], estilo[0], estilo[1])
            
    
    def grafico_linha(self, celula: str, comprimento: float, altura: float, 
                      titulo: str,
                      titulo_x: str, titulo_y: str,
                      referencia_x: Reference,
                      referencia_y: Reference,
                      propriedades_grafico: list):
        grafico = LineChart()
        grafico.width = comprimento
        grafico.height = altura
        grafico.title = titulo
        grafico.x_axis.title = titulo_x
        grafico.y_axis.title = titulo_y
        
        grafico.add_data(referencia_x)
        grafico.set_categories(referencia_y)
        
        for serie, propriedade in zip(grafico.series, propriedades_grafico):
            serie.graphicalProperties.line.width = propriedade.espessura
            serie.graphicalProperties.line.solidFill = propriedade.cor
            
        self.ws_ativa.add_chart(grafico, celula)
        
        
    def adiciona_imagem(self, celula: str, caminho_imagem: str):
        imagem = Image(caminho_imagem)
        self.ws_ativa.add_image(imagem, celula)
        
        
    def salva_arquivo(self, caminho: str):
        self.workbook.save(caminho)